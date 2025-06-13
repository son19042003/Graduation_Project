import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
FILE_PATH = os.path.join(BASE_DIR, "data/testcases/testcases.xlsx")

def read_test_cases(sheet_name, filepath=FILE_PATH):
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        raise FileNotFoundError(f"Không tìm thấy file test case tại: {filepath}")
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' không được tìm thấy trong file test case.")
    
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=False))
    if not rows or len(rows) < 2:
        raise ValueError(f"Sheet '{sheet_name}' không có dữ liệu hợp lệ.")
    
    headers = [cell.value for cell in rows[0]]
    data = []

    for i, row in enumerate(rows[1:], start=2):
        if all(cell.value is None for cell in row):
            continue
        test_case = {headers[j]: cell.value for j, cell in enumerate(row)}
        test_case['_row'] = i
        data.append(test_case)
    
    wb.close()
    return data

def write_test_result(sheet_name, row_number, actual, result, filepath=FILE_PATH):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]

    headers = [cell.value for cell in sheet[1]]
    try:
        actual_col = headers.index('Thực tế') + 1
        result_col = headers.index('Kết quả') + 1
    except ValueError:
        raise ValueError(f"File Excel phải có các cột 'Thực tế' và 'Kết quả'")
    
    sheet.cell(row=row_number, column=actual_col, value=actual)
    sheet.cell(row=row_number, column=result_col, value=result)

    wb.save(filepath)
    wb.close()