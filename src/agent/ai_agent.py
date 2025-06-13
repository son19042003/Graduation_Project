import requests
import os
import json
from openpyxl import Workbook, load_workbook
import argparse

API_KEY = "AIzaSyBKQ5J3HglHv-OhhkQei7Bo_bpZ8i1S5pM"
API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={API_KEY}"

HEADERS = {
    "Content-Type": "application/json"
}

PROMPT_DIR = os.path.join(os.path.dirname(__file__), 'prompts')
TEMPLATE_PATH = os.path.join(PROMPT_DIR, 'template.txt')
EXCEL_OUTPUT_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../data/testcases/testcases.xlsx'))
PROCESSED_TRACK_PATH = os.path.join(PROMPT_DIR, 'processed_prompts.json')

def load_template():
    if not os.path.exists(TEMPLATE_PATH):
        print(f'[ERROR] Thiếu tệp mẫu tại: {TEMPLATE_PATH}')
        return ""
    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        return f.read().strip()

def generate_testcase_from_prompt(full_prompt: str) -> str:
    payload = {
        "contents": [{"parts": [{'text': full_prompt}]}]
    }
    try:
        response = requests.post(API_URL, headers=HEADERS, json=payload)
        response.raise_for_status()
        data = response.json()
        return data['candidates'][0]['content']['parts'][0]['text']
    except requests.exceptions.RequestException as e:
        print(f'[ERROR] Gửi yêu cầu thất bại: {e}')
        return ""

def parse_testcase_text(text: str):
    rows = []
    buffer = ""
    lines = text.strip().splitlines()

    cleaned_lines = [
        line for line in lines
        if not line.strip().startswith('```') and line.strip() != '```'
    ]

    if cleaned_lines and cleaned_lines[0].strip().lower().startswith('id;'):
        cleaned_lines = cleaned_lines[1:]
    
    for line in cleaned_lines:
        if not line.strip():
            continue

        buffer += ' ' + line.strip()
        semicolon_count = buffer.count(';')

        if semicolon_count >= 7:
            parts = [col.strip() for col in buffer.split(';')]
            if len(parts) < 8:
                parts += [''] * (8 - len(parts))
            elif len(parts) > 8:
                parts = parts[:8]
            rows.append(parts)
            buffer = ""
    return rows

def load_processed_prompts():
    if not os.path.exists(PROCESSED_TRACK_PATH):
        return []
    with open(PROCESSED_TRACK_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_processed_prompts(processed):
    with open(PROCESSED_TRACK_PATH, 'w', encoding='utf-8') as f:
        json.dump(processed, f, ensure_ascii=False, indent=2)

def main():
    parser = argparse.ArgumentParser(description="Tạo test cases từ prompt file")
    parser.add_argument('--reprocess', nargs='*', help="Danh sách các file prompt muốn chạy lại")
    args = parser.parse_args()

    reprocess_list = set(args.reprocess) if args.reprocess else set()

    if not os.path.exists(PROMPT_DIR):
        print(f'[ERROR] Không tìm thấy thư mục prompt: {PROMPT_DIR}')
        return
    
    template_text = load_template()
    if not  template_text:
        return
    
    processed = set(load_processed_prompts())
    
    if os.path.exists(EXCEL_OUTPUT_PATH):
        wb = load_workbook(EXCEL_OUTPUT_PATH)
    else:
        wb = Workbook()

    default_sheet = wb.active if wb.active.title =='Sheet' and len(wb.sheetnames) == 1 else None

    prompts = [f for f in os.listdir(PROMPT_DIR) if f.endswith('.txt') and f != 'template.txt']
    new_processed = set(processed)
    any_sheet_created = False
    
    for file in prompts:
        func_name = os.path.splitext(file)[0].capitalize()
        if file in processed and file not in reprocess_list:
            print(f"[SKIP] '{file}' đã xử lý, bỏ qua.")
            continue

        prompt_path = os.path.join(PROMPT_DIR, file)
        with open(prompt_path, 'r', encoding='utf-8') as f:
            prompt_text = f.read().strip()

        combined_prompt = f"{prompt_text.strip()}\n\n{template_text}"

        print(f"\n[INFO] Đang tạo test cases cho: '{func_name}'...")
        generated_text = generate_testcase_from_prompt(combined_prompt)

        print(f"\n--- Nội dung được AI sinh ra từ prompt {file} ---\n")
        print(generated_text)
        print("\n--- Kết thúc nội dung sinh ---\n")

        testcases = parse_testcase_text(generated_text)

        if not testcases:
            print(f"[WARN] Không có test case nào được tạo từ: {file}")
            continue

        ws = wb.create_sheet(title=func_name)
        headers = ["ID", "Chức năng", "Loại test case", "Mô tả", "Dữ liệu test", "Kỳ vọng", "Thực tế", "Kết quả"]
        ws.append(headers)

        for row in testcases:
            ws.append(row)

        print(f"[DONE] Đã tạo sheet: {func_name}")
        new_processed.add(file)
        any_sheet_created = True

    if default_sheet and any_sheet_created:
        wb.remove(default_sheet)
    
    if any_sheet_created:
        os.makedirs(os.path.dirname(EXCEL_OUTPUT_PATH), exist_ok=True)
        wb.save(EXCEL_OUTPUT_PATH)
        save_processed_prompts(list(new_processed))
        print(f"\nTất cả test cases mới đã được lưu tại: {EXCEL_OUTPUT_PATH}")
    else:
        print(f"\nKhông có test case nào được tạo, file Excel không thay đổi!")

if __name__ == "__main__":
    main()